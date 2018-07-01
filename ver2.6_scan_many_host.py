#!/usr/bin/env python
#TheZero
#This code is under Public Domain

from threading import Thread
import socket
#host = input('host > ')
from_port = 0
to_port = 1024
counting_open = []
counting_close = []
threads = []

from openpyxl import load_workbook
# set file path
filepath="IP.xlsx"
# load demo.xlsx 
wb=load_workbook('IP.xlsx')
# activate demo.xlsx
sheet=wb.active
# get b1 cell value
#b1=sheet['A'+'1'].value
# get b2 cell value
#b2=sheet['A2'].value
row=sheet.max_row
print(row)
print(' IP duoc scan')
row=sheet.max_row + 1

for i in range(1, row):
	i=str(i)
	host=sheet['A'+i].value
	def scan(port):
		s = socket.socket()
		result = s.connect_ex((host,port))
		#print('working on port > '+(str(port)))      
		if result == 0:
			counting_open.append(port)
			#print((str(port))+' -> open')
			#sheet['F'+i]=counting_open
			s.close()
		else:
			counting_close.append(port)
			#print((str(port))+' -> close') 
			s.close()
	#wb.save("IP.xlsx")

	for i in range(from_port, to_port+1):
		t = Thread(target=scan, args=(i,))
		threads.append(t)
		t.start()
	
	[x.join() for x in threads]
	print(counting_open)
	temp=chr(ord('A') + 1)
	temp=temp+i
	#print(temp)
	counting_open=str(counting_open)
	sheet[temp]=counting_open
	counting_open = []
wb.save("IP.xlsx")