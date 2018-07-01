# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="IP.xlsx"
# load demo.xlsx 
wb=load_workbook('IP.xlsx')
# activate demo.xlsx
sheet=wb.active
# get b1 cell value
#b1=sheet['A'+'1'].value
# get b2 cell value
#b2=sheet['A2'].value
row=sheet.max_row + 1
print(row)
for i in range(1, row):
	i=str(i)
	temp=sheet['A'+i].value
	print(temp)
for i in range(1, row):
	i=str(i)
	sheet['F'+i]='done'
wb.save("IP.xlsx")
#b3=sheet.max_row().value
# get b3 cell value
# print b1, b2 and b3